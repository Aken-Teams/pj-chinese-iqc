"""Parse PJXZ review rules Excel and fuzzy-match param names to DB data."""

from __future__ import annotations

import re
from typing import Any

import openpyxl
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.models.product import Product
from app.models.vendor import Vendor
from app.models.review import ReviewRule


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def parse_review_rules_excel(file_path: str) -> dict[str, Any]:
    """Parse the PJXZ review rules Excel file.

    Returns dict with keys: sheets_parsed, rules (list of raw rule dicts).
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    all_rules: list[dict] = []
    sheets_parsed: list[str] = []

    for sheet_name in wb.sheetnames:
        code = sheet_name.strip().upper()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue
        sheets_parsed.append(code)
        all_rules.extend(_parse_sheet(rows, code))

    wb.close()
    return {"sheets_parsed": sheets_parsed, "rules": all_rules}


def _parse_sheet(rows: list[tuple], vendor_code: str) -> list[dict]:
    """Parse a single sheet from the review rules Excel."""
    row2 = rows[1]  # param names
    row3 = rows[2]  # Q1, Q2, Q3 labels

    # Detect param positions by scanning row2 from col index 3 (col D)
    param_columns: list[dict] = []
    col = 3
    max_col = len(row2)
    while col < max_col:
        val = row2[col]
        if val is not None and str(val).strip() and str(val).strip() not in ("Q1", "Q2", "Q3", "L LIMIT", "U LIMIT"):
            param_name = str(val).strip()
            # Count Q-levels by scanning row3 for sequential Q1, Q2, Q3
            q_count = 0
            expected_q = 1
            for qc in range(col, min(col + 18, len(row3) if row3 else 0), 2):
                q_label = str(row3[qc]).strip().upper() if row3[qc] else ""
                if q_label == f"Q{expected_q}":
                    q_count += 1
                    expected_q += 1
                else:
                    break
            if q_count == 0:
                q_count = 2  # default
            param_columns.append({
                "name": param_name,
                "col_start": col,
                "q_count": q_count,
            })
            col += q_count * 2
        else:
            col += 1

    # Parse data rows (row 5+, index 4+)
    rules: list[dict] = []
    for row_idx in range(4, len(rows)):
        row = rows[row_idx]
        if not row or not row[0]:
            continue

        product_code_pjxz = str(row[0]).strip() if row[0] else ""
        product_id_vendor = str(row[1]).strip() if len(row) > 1 and row[1] else ""

        if not product_id_vendor:
            # Alias row (no product ID) — skip
            continue

        for param_info in param_columns:
            cs = param_info["col_start"]
            qc = param_info["q_count"]

            limits: dict[str, float | None] = {}
            has_any = False
            for qi in range(qc):
                base = cs + qi * 2
                lo = _parse_limit(row[base] if base < len(row) else None)
                hi = _parse_limit(row[base + 1] if base + 1 < len(row) else None)
                limits[f"q{qi + 1}_lower"] = lo
                limits[f"q{qi + 1}_upper"] = hi
                if lo is not None or hi is not None:
                    has_any = True

            # Fill missing Q-levels with None
            for qi in range(qc, 3):
                limits[f"q{qi + 1}_lower"] = None
                limits[f"q{qi + 1}_upper"] = None

            if not has_any:
                continue

            rules.append({
                "product_code": product_id_vendor,
                "product_code_pjxz": product_code_pjxz,
                "vendor_code": vendor_code,
                "param_name_excel": param_info["name"],
                **limits,
            })

    return rules


def _parse_limit(val: Any) -> float | None:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("/", "", "-", "N/A", "n/a"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Fuzzy matching
# ---------------------------------------------------------------------------

def _normalize_param_name(name: str) -> str:
    """Normalize for fuzzy comparison:
    - lowercase
    - strip digit between alpha prefix and _ (rds1_ -> rds_)
    - strip leading minus after _ (_-250 -> _250)
    """
    s = name.strip().lower()
    s = re.sub(r"^([a-z]+)\d+_", r"\1_", s)
    s = re.sub(r"_-(\d)", r"_\1", s)
    return s


def _levenshtein(s1: str, s2: str) -> int:
    if len(s1) < len(s2):
        return _levenshtein(s2, s1)
    if len(s2) == 0:
        return len(s1)
    prev = list(range(len(s2) + 1))
    for i, c1 in enumerate(s1):
        curr = [i + 1]
        for j, c2 in enumerate(s2):
            curr.append(min(prev[j + 1] + 1, curr[j] + 1, prev[j] + (c1 != c2)))
        prev = curr
    return prev[-1]


def fuzzy_match_param(
    excel_name: str,
    db_params: list[str],
) -> tuple[str | None, float]:
    """Return (best_match, confidence) for an Excel param against DB params."""
    if not db_params:
        return None, 0.0

    # 1) Exact (case-insensitive)
    el = excel_name.strip().lower()
    for p in db_params:
        if p.strip().lower() == el:
            return p, 1.0

    # 2) Normalized exact
    en = _normalize_param_name(excel_name)
    for p in db_params:
        if _normalize_param_name(p) == en:
            return p, 0.95

    # 3) Normalized containment
    best, best_score = None, 0.0
    for p in db_params:
        pn = _normalize_param_name(p)
        if en in pn or pn in en:
            if 0.8 > best_score:
                best_score = 0.8
                best = p

    # 4) Levenshtein on normalized
    if best_score < 0.8:
        for p in db_params:
            pn = _normalize_param_name(p)
            d = _levenshtein(en, pn)
            mx = max(len(en), len(pn))
            if mx == 0:
                continue
            sim = (1.0 - d / mx) * 0.9
            if sim > best_score:
                best_score = sim
                best = p

    if best_score >= 0.6:
        return best, round(best_score, 2)
    return None, 0.0


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def get_product_param_names(db: Session, product_id: int) -> list[str]:
    """Return distinct electrical param names for a product."""
    result = db.execute(
        text("""
            SELECT DISTINCT ev.param_name
            FROM electrical_values ev
            JOIN die_data dd ON ev.die_id = dd.id
            JOIN wafers w ON dd.wafer_id = w.id
            JOIN lots l ON w.lot_id = l.id
            WHERE l.product_id = :pid
            ORDER BY ev.param_name
        """),
        {"pid": product_id},
    ).fetchall()
    return [r[0] for r in result]


def build_import_preview(
    db: Session,
    parsed_rules: list[dict],
) -> list[dict]:
    """Enrich parsed rules with DB product matching and param fuzzy matching.

    Returns list of preview row dicts with status field.
    """
    # Cache: vendor_code -> vendor_id
    vendors = {v.code: v.id for v in db.query(Vendor).all()}
    # Cache: (vendor_id, product_code) -> product
    product_cache: dict[tuple[int, str], Product | None] = {}
    # Cache: product_id -> param names
    param_cache: dict[int, list[str]] = {}
    # Cache: product_id -> existing rules
    existing_cache: dict[int, dict[str, ReviewRule]] = {}

    preview: list[dict] = []
    for rule in parsed_rules:
        vendor_id = vendors.get(rule["vendor_code"])
        product_code = rule["product_code"]

        # Lookup product
        cache_key = (vendor_id or 0, product_code)
        if cache_key not in product_cache:
            q = db.query(Product).filter(Product.product_code == product_code)
            if vendor_id:
                q = q.filter(Product.vendor_id == vendor_id)
            product_cache[cache_key] = q.first()
        product = product_cache[cache_key]

        if not product:
            # Vendor known → can auto-create the product on confirm.
            # Vendor unknown → real error, cannot import.
            preview.append({
                **rule,
                "product_id": None,
                "param_name_matched": rule["param_name_excel"] if vendor_id else None,
                "match_confidence": 0.0,
                "status": "auto_create" if vendor_id else "no_vendor",
            })
            continue

        # Get DB param names
        if product.id not in param_cache:
            param_cache[product.id] = get_product_param_names(db, product.id)
        db_params = param_cache[product.id]

        # Fuzzy match
        matched, confidence = fuzzy_match_param(rule["param_name_excel"], db_params)

        # Check existing rule
        if product.id not in existing_cache:
            existing_rules = (
                db.query(ReviewRule)
                .filter(ReviewRule.product_id == product.id)
                .all()
            )
            existing_cache[product.id] = {r.param_name: r for r in existing_rules}

        final_param = matched or rule["param_name_excel"]
        has_existing = final_param in existing_cache[product.id]

        if matched:
            status = "update_existing" if has_existing else "ready"
        else:
            status = "no_param_match"

        preview.append({
            **rule,
            "product_id": product.id,
            "param_name_matched": matched,
            "match_confidence": confidence,
            "status": status,
        })

    return preview
