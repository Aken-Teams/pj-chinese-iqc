"""One cell-grid abstraction over every CP file shape we receive.

The 2026-08 survey (無錫 six vendors + 徐州) turned up .xlsx, .xlsm, .xls, .csv,
comma- and tab-delimited .txt — and metadata that sits at a fixed cell in one
vendor's files but drifts by several rows in another's. Rather than teach each
parser those variations, everything is loaded into a `Grid` that exposes
1-indexed cell access plus label-anchored lookup.
"""
from __future__ import annotations

import csv
import io
import os
import re
from typing import Any, Iterator, Optional

# Encodings to try, in order. CP dumps are mostly ASCII but can carry a stray
# GBK/Big5 byte (a Ω unit symbol, a Chinese operator name).
_ENCODINGS = ("utf-8-sig", "utf-8", "gbk", "big5", "latin-1")


class Grid:
    """A 1-indexed rectangular view of a spreadsheet or delimited text file."""

    def __init__(self, rows: list[list[Any]], sheets: list[str],
                 sheet_used: str, encoding: str | None, delimiter: str | None):
        self._rows = rows
        self.sheets = sheets
        self.sheet_used = sheet_used
        self.encoding = encoding
        self.delimiter = delimiter

    # --- dimensions ------------------------------------------------------
    @property
    def n_rows(self) -> int:
        return len(self._rows)

    @property
    def n_cols(self) -> int:
        return max((len(r) for r in self._rows), default=0)

    # --- access ----------------------------------------------------------
    def cell(self, row: int, col: int) -> Any:
        """1-indexed cell value. Blank strings normalise to None so every
        caller sees 'empty' the same way regardless of source format."""
        if row < 1 or col < 1 or row > len(self._rows):
            return None
        r = self._rows[row - 1]
        if col > len(r):
            return None
        v = r[col - 1]
        if isinstance(v, str):
            v = v.strip()
            return v or None
        return v

    def row(self, row: int) -> list[Any]:
        if row < 1 or row > len(self._rows):
            return []
        return list(self._rows[row - 1])

    def iter_rows(self, start_row: int = 1) -> Iterator[tuple[int, list[Any]]]:
        for i in range(start_row, len(self._rows) + 1):
            yield i, self._rows[i - 1]

    def head(self, n: int = 40, cols: int = 20) -> list[list[Any]]:
        """Top-left block, for previews and for the layout detector."""
        return [[self.cell(r, c) for c in range(1, cols + 1)]
                for r in range(1, min(n, len(self._rows)) + 1)]

    # --- label-anchored lookup -------------------------------------------
    def find_label(self, text: str, search_rows: int = 80,
                   search_cols: int = 12) -> Optional[tuple[int, int]]:
        """Locate a label cell, case- and punctuation-insensitively.

        Vendors write the same label as "LOT ID:", "Lot number", "批号" —
        matching is loose (normalised substring, either direction) because the
        alternative is a brittle exact match that breaks on a trailing colon.
        """
        needle = _norm(text)
        if not needle:
            return None
        limit = min(search_rows, len(self._rows))

        # Two passes so an exact label always beats a partial one, and so a
        # short cell can never swallow the match: matching in the "cell is
        # contained by the search text" direction let a lone "A" satisfy a
        # search for "Wafer ID" (both normalise into 'waferid'), which returned
        # a neighbouring row's value. Only "cell contains the search text" is
        # allowed now, which still finds "LOT ID:" for "lot id".
        for exact in (True, False):
            for r in range(1, limit + 1):
                for c in range(1, search_cols + 1):
                    v = self.cell(r, c)
                    if not isinstance(v, str):
                        continue
                    hay = _norm(v)
                    if not hay:
                        continue
                    if exact:
                        if hay == needle:
                            return (r, c)
                    elif needle in hay:
                        return (r, c)
        return None

    def label_value(self, text: str, search_rows: int = 80) -> Any:
        """Value belonging to a label: the first non-empty cell to its right,
        falling back to the cell directly below.

        This is what makes 禾纳 parseable — its `wafer number` row moves
        between files because the Bias block above it varies in length, so a
        fixed "row,col" address cannot address it.
        """
        pos = self.find_label(text, search_rows=search_rows)
        if pos is None:
            return None
        r, c = pos
        for cc in range(c + 1, min(c + 8, self.n_cols + 1)):
            v = self.cell(r, cc)
            if v is not None:
                return v
        return self.cell(r + 1, c)


def _norm(s: str) -> str:
    """Fold case, spaces and separator punctuation for label matching."""
    return re.sub(r"[\s:：#\-_.,，、]+", "", str(s)).lower()


def parse_cell_ref(ref: str | None) -> Optional[tuple[int, int]]:
    """Parse a cell reference into (row, col), both 1-indexed.

    Accepts the stored "row,col" form ("4,2") and spreadsheet notation
    ("B4"), since the template UI lets people type either.
    """
    if not ref:
        return None
    s = str(ref).strip()
    if not s:
        return None
    if "," in s:
        parts = s.split(",")
        if len(parts) != 2:
            return None
        try:
            r, c = int(parts[0].strip()), int(parts[1].strip())
        except (TypeError, ValueError):
            return None
        return (r, c) if r >= 1 and c >= 1 else None
    m = re.fullmatch(r"([A-Za-z]{1,3})\s*(\d+)", s)
    if not m:
        return None
    letters, digits = m.group(1).upper(), m.group(2)
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    row = int(digits)
    return (row, col) if row >= 1 else None


def _sniff_delimiter(text: str) -> str:
    probe = text[:16384]
    return "\t" if probe.count("\t") > probe.count(",") else ","


def _decode(raw: bytes) -> tuple[str, str]:
    for enc in _ENCODINGS:
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace"), "latin-1"


def _pick_sheet(names: list[str], selector: str | None) -> str:
    """Resolve a sheet selector. "#2" selects by 1-indexed position; a plain
    string selects by name (case-insensitive). Falling back to a sheet called
    "data", then the first — the historical behaviour."""
    if selector:
        s = selector.strip()
        if s.startswith("#"):
            try:
                idx = int(s[1:]) - 1
            except ValueError:
                idx = -1
            if 0 <= idx < len(names):
                return names[idx]
        for n in names:
            if n.lower() == s.lower():
                return n
    for n in names:
        if n.lower() == "data":
            return n
    return names[0]


def open_grid(path: str, sheet_selector: str | None = None,
              delimiter: str | None = None, max_rows: int | None = None) -> Grid:
    """Load any supported CP file into a Grid.

    `delimiter` is "tab"/"comma" to force one, or None to sniff.
    `max_rows` caps the load for preview/detection work.
    """
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        import openpyxl

        def _read(read_only: bool) -> tuple[list[list[Any]], list[str], str]:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=read_only)
            try:
                name = _pick_sheet(wb.sheetnames, sheet_selector)
                ws = wb[name]
                # Read-only mode trusts the workbook's stored <dimension>.
                # Some writers emit "A1" regardless of content, which made a
                # real 280 KB vendor file look like a single empty cell.
                if read_only and getattr(ws, "reset_dimensions", None):
                    if (ws.max_row or 0) <= 1 and (ws.max_column or 0) <= 1:
                        ws.reset_dimensions()
                out: list[list[Any]] = []
                for i, r in enumerate(ws.iter_rows(values_only=True)):
                    if max_rows is not None and i >= max_rows:
                        break
                    out.append(list(r))
                return out, list(wb.sheetnames), name
            finally:
                wb.close()

        rows, sheets, name = _read(True)
        if not any(any(v is not None for v in r) for r in rows):
            # Still nothing: the streaming reader cannot recover this file.
            # Fall back to a full parse, which rebuilds dimensions itself.
            rows, sheets, name = _read(False)
        return Grid(rows, sheets, name, None, None)

    if ext == ".xls":
        # Legacy BIFF; pandas+xlrd reads it without a conversion round-trip.
        import pandas as pd
        sheets = pd.read_excel(path, sheet_name=None, header=None, engine="xlrd")
        names = list(sheets)
        name = _pick_sheet(names, sheet_selector)
        df = sheets[name]
        if max_rows is not None:
            df = df.head(max_rows)
        rows = df.where(df.notna(), None).values.tolist()
        return Grid(rows, names, name, None, None)

    # .csv / .txt / anything delimited
    with open(path, "rb") as f:
        raw = f.read()
    text, enc = _decode(raw)
    if delimiter == "tab":
        delim = "\t"
    elif delimiter == "comma":
        delim = ","
    else:
        delim = _sniff_delimiter(text)
    rows = []
    for i, r in enumerate(csv.reader(io.StringIO(text), delimiter=delim)):
        if max_rows is not None and i >= max_rows:
            break
        rows.append(r)
    return Grid(rows, ["-"], "-", enc, "tab" if delim == "\t" else "comma")


SUPPORTED_EXTENSIONS = (".xlsx", ".xlsm", ".xls", ".csv", ".txt")


def is_supported(filename: str) -> bool:
    return os.path.splitext(filename)[1].lower() in SUPPORTED_EXTENSIONS
