import openpyxl
from .base import BaseParser, ParseResult, ParsedWafer, ParsedDie, ParsedCpSpec


class JJWParser(BaseParser):
    HEADER_ROW = 8
    DATA_START_ROW = 9
    LOWER_LIMIT_ROW = 2
    UPPER_LIMIT_ROW = 3
    ELECTRICAL_START_COL = 15  # column O
    WAFER_ID_COL = 4           # column D
    BIN_COL = 12               # column L
    X_COORD_COL = 13           # column M
    Y_COORD_COL = 14           # column N
    PRODUCT_ID_COL = 1         # column A
    LOT_ID_COL = 2             # column B

    def _open_workbook(self, filepath: str):
        return openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    def _get_data_sheet(self, wb):
        # Try "data" sheet first, then first sheet
        if "data" in wb.sheetnames:
            return wb["data"]
        return wb[wb.sheetnames[0]]

    def preview(self, filepath: str) -> dict:
        wb = self._open_workbook(filepath)
        ws = self._get_data_sheet(wb)

        # Read param names from header row
        param_names = []
        col = self.ELECTRICAL_START_COL
        while True:
            val = ws.cell(row=self.HEADER_ROW, column=col).value
            if val is None:
                break
            param_names.append(str(val).strip())
            col += 1

        # Count data rows and detect wafers
        wafer_ids = set()
        row_count = 0
        row = self.DATA_START_ROW
        for r in ws.iter_rows(min_row=self.DATA_START_ROW, max_col=self.WAFER_ID_COL, values_only=False):
            cell_val = r[self.WAFER_ID_COL - 1].value
            if cell_val is None:
                break
            wafer_ids.add(str(cell_val))
            row_count += 1

        product_id = ws.cell(row=self.DATA_START_ROW, column=self.PRODUCT_ID_COL).value
        lot_id = ws.cell(row=self.DATA_START_ROW, column=self.LOT_ID_COL).value

        wb.close()
        return {
            "wafersDetected": len(wafer_ids),
            "diePerWafer": None,  # variable
            "dataRows": row_count,
            "format": "JJW",
            "productId": str(product_id) if product_id else None,
            "lotId": str(lot_id) if lot_id else None,
            "paramNames": param_names,
        }

    def parse(self, filepath: str) -> ParseResult:
        wb = self._open_workbook(filepath)
        ws = self._get_data_sheet(wb)

        # 1. Read param names from header row
        param_names = []
        col = self.ELECTRICAL_START_COL
        while True:
            val = ws.cell(row=self.HEADER_ROW, column=col).value
            if val is None:
                break
            param_names.append(str(val).strip())
            col += 1

        # 2. Read CP specs from limit rows
        cp_specs = []
        for i, pname in enumerate(param_names):
            ecol = self.ELECTRICAL_START_COL + i
            lower = ws.cell(row=self.LOWER_LIMIT_ROW, column=ecol).value
            upper = ws.cell(row=self.UPPER_LIMIT_ROW, column=ecol).value
            cp_specs.append(ParsedCpSpec(
                param_name=pname,
                lower_limit=float(lower) if lower is not None and lower != "" else None,
                upper_limit=float(upper) if upper is not None and upper != "" else None,
            ))

        # 3. Read die data, group by wafer
        wafers_dict: dict[str, list[ParsedDie]] = {}
        product_id = None
        lot_id = None
        mark_lot_id = None
        test_program = None
        total_rows = 0

        for row in ws.iter_rows(min_row=self.DATA_START_ROW, values_only=False):
            wafer_val = row[self.WAFER_ID_COL - 1].value
            if wafer_val is None:
                break

            total_rows += 1
            wid = str(wafer_val).strip()

            if product_id is None:
                product_id = str(row[self.PRODUCT_ID_COL - 1].value or "")
                lot_id = str(row[self.LOT_ID_COL - 1].value or "")
                mark_lot_id = str(row[2].value or "") if len(row) > 2 else None  # col C
                test_program = str(row[4].value or "") if len(row) > 4 else None  # col E

            bin_val = row[self.BIN_COL - 1].value
            x_val = row[self.X_COORD_COL - 1].value if len(row) >= self.X_COORD_COL else None
            y_val = row[self.Y_COORD_COL - 1].value if len(row) >= self.Y_COORD_COL else None

            electrical = {}
            for j, pname in enumerate(param_names):
                ecol_idx = self.ELECTRICAL_START_COL - 1 + j
                if ecol_idx < len(row):
                    v = row[ecol_idx].value
                    electrical[pname] = float(v) if v is not None else None
                else:
                    electrical[pname] = None

            die = ParsedDie(
                site_no=None,
                bin=int(bin_val) if bin_val is not None else 0,
                x_coord=int(x_val) if x_val is not None else None,
                y_coord=int(y_val) if y_val is not None else None,
                electrical=electrical,
            )

            if wid not in wafers_dict:
                wafers_dict[wid] = []
            wafers_dict[wid].append(die)

        wb.close()

        # Build wafer objects
        wafers = []
        for wid, dies in wafers_dict.items():
            bin1_count = sum(1 for d in dies if d.bin == 1)
            wafers.append(ParsedWafer(
                wafer_id=wid,
                gross_die=len(dies),
                bin1_count=bin1_count,
                dies=dies,
            ))

        return ParseResult(
            product_id=product_id or "",
            lot_id=lot_id or "",
            mark_lot_id=mark_lot_id,
            test_program=test_program,
            vendor_code="JJW",
            wafers=wafers,
            cp_specs=cp_specs,
            param_names=param_names,
            total_rows=total_rows,
        )
