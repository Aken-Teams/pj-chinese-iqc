"""
XRW (祥瑞微) CP data parser — supports two file layouts:

**Legacy format** (has a "data" sheet with fixed column positions):
  Row 2: LOWER LIMIT, Row 3: UPPER LIMIT, Row 5: Header, Row 6+: Data
  Columns: ProductID(A) LotID(B) ... WaferID(D) ... BIN(I) X(J) Y(K) Params(L+)

**New per-wafer format** (single sheet, metadata header):
  Auto-detected by scanning for "Parameter" / "Max" / "Min" markers.
  Wafer ID in the last column ("ID"), electrical params between X/Y and ID.
"""

import openpyxl
from .base import BaseParser, ParseResult, ParsedWafer, ParsedDie, ParsedCpSpec


class XRWParser(BaseParser):

    # ------------------------------------------------------------------
    # Legacy format constants  (PPT sample / "data" sheet)
    # ------------------------------------------------------------------
    _L_HEADER_ROW = 5
    _L_DATA_START_ROW = 6
    _L_LOWER_LIMIT_ROW = 2
    _L_UPPER_LIMIT_ROW = 3
    _L_ELECTRICAL_START_COL = 12  # column L
    _L_WAFER_ID_COL = 4           # column D
    _L_BIN_COL = 9                # column I
    _L_X_COORD_COL = 10           # column J
    _L_Y_COORD_COL = 11           # column K
    _L_PRODUCT_ID_COL = 1
    _L_LOT_ID_COL = 2
    _L_FIXED_DIE_COUNT = 208

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _open_workbook(self, filepath: str):
        return openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    @staticmethod
    def _safe_float(val):
        if val is None or val == "":
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def _is_legacy_format(self, wb) -> bool:
        """Legacy format has a 'data' sheet with 'ProductID' in row 5 col A."""
        if "data" not in wb.sheetnames:
            return False
        ws = wb["data"]
        val = ws.cell(row=self._L_HEADER_ROW, column=1).value
        return val is not None and str(val).strip() == "ProductID"

    def _get_sheet(self, wb, legacy: bool):
        if legacy:
            return wb["data"]
        return wb[wb.sheetnames[0]]

    # ==================================================================
    # LEGACY FORMAT  (PPT sample with "data" sheet)
    # ==================================================================

    def _legacy_preview(self, ws) -> dict:
        param_names = []
        col = self._L_ELECTRICAL_START_COL
        while True:
            val = ws.cell(row=self._L_HEADER_ROW, column=col).value
            if val is None:
                break
            param_names.append(str(val).strip())
            col += 1

        wafer_ids: set[str] = set()
        row_count = 0
        for r in ws.iter_rows(min_row=self._L_DATA_START_ROW,
                              max_col=self._L_WAFER_ID_COL, values_only=False):
            cell_val = r[self._L_WAFER_ID_COL - 1].value
            if cell_val is None:
                break
            wafer_ids.add(str(cell_val))
            row_count += 1

        product_id = ws.cell(row=self._L_DATA_START_ROW,
                             column=self._L_PRODUCT_ID_COL).value
        lot_id = ws.cell(row=self._L_DATA_START_ROW,
                         column=self._L_LOT_ID_COL).value

        return {
            "wafersDetected": len(wafer_ids),
            "diePerWafer": self._L_FIXED_DIE_COUNT,
            "dataRows": row_count,
            "format": "XRW",
            "productId": str(product_id) if product_id else None,
            "lotId": str(lot_id) if lot_id else None,
            "paramNames": param_names,
        }

    def _legacy_parse(self, ws) -> ParseResult:
        # Param names
        param_names: list[str] = []
        col = self._L_ELECTRICAL_START_COL
        while True:
            val = ws.cell(row=self._L_HEADER_ROW, column=col).value
            if val is None:
                break
            param_names.append(str(val).strip())
            col += 1

        # CP specs
        cp_specs: list[ParsedCpSpec] = []
        for i, pname in enumerate(param_names):
            ecol = self._L_ELECTRICAL_START_COL + i
            lower = ws.cell(row=self._L_LOWER_LIMIT_ROW, column=ecol).value
            upper = ws.cell(row=self._L_UPPER_LIMIT_ROW, column=ecol).value
            cp_specs.append(ParsedCpSpec(
                param_name=pname,
                lower_limit=self._safe_float(lower),
                upper_limit=self._safe_float(upper),
            ))

        # Die data
        wafers_dict: dict[str, list[ParsedDie]] = {}
        product_id = lot_id = mark_lot_id = test_program = None
        total_rows = 0

        for row in ws.iter_rows(min_row=self._L_DATA_START_ROW, values_only=False):
            wafer_val = row[self._L_WAFER_ID_COL - 1].value
            if wafer_val is None:
                break
            total_rows += 1
            wid = str(wafer_val).strip()

            if product_id is None:
                product_id = str(row[self._L_PRODUCT_ID_COL - 1].value or "")
                lot_id = str(row[self._L_LOT_ID_COL - 1].value or "")
                mark_lot_id = str(row[2].value or "") if len(row) > 2 else None
                test_program = str(row[4].value or "") if len(row) > 4 else None

            bin_val = row[self._L_BIN_COL - 1].value
            x_val = row[self._L_X_COORD_COL - 1].value
            y_val = row[self._L_Y_COORD_COL - 1].value

            electrical: dict[str, float | None] = {}
            for j, pname in enumerate(param_names):
                idx = self._L_ELECTRICAL_START_COL - 1 + j
                if idx < len(row):
                    electrical[pname] = self._safe_float(row[idx].value)
                else:
                    electrical[pname] = None

            die = ParsedDie(
                site_no=None,
                bin=int(bin_val) if bin_val is not None else 0,
                x_coord=int(x_val) if x_val is not None else None,
                y_coord=int(y_val) if y_val is not None else None,
                electrical=electrical,
            )
            wafers_dict.setdefault(wid, []).append(die)

        wafers = []
        for wid, dies in wafers_dict.items():
            bin1_count = sum(1 for d in dies if d.bin == 1)
            wafers.append(ParsedWafer(
                wafer_id=wid,
                gross_die=len(dies),
                bin1_count=bin1_count,
                dies=dies,
            ))

        return ParseResult(
            product_id=product_id or "",
            lot_id=lot_id or "",
            mark_lot_id=mark_lot_id,
            test_program=test_program,
            vendor_code="XRW",
            wafers=wafers,
            cp_specs=cp_specs,
            param_names=param_names,
            total_rows=total_rows,
        )

    # ==================================================================
    # NEW PER-WAFER FORMAT  (H2XN15-style, auto-detected)
    # ==================================================================

    def _scan_structure(self, ws):
        """Forward scan to detect row positions and metadata."""
        meta: dict[str, str] = {}
        header_row = None
        upper_limit_row = None
        lower_limit_row = None
        data_start_row = None

        for i, cells in enumerate(ws.iter_rows(min_row=1, max_col=6,
                                               values_only=True)):
            rn = i + 1
            a = str(cells[0]).strip() if cells[0] is not None else ""

            if a.startswith("LOT ID"):
                meta["lot_id"] = str(cells[1]).strip() if cells[1] else ""
            elif a.startswith("Device Name"):
                meta["device_name"] = str(cells[1]).strip() if cells[1] else ""
            elif a.startswith("Test Program"):
                meta["test_program"] = str(cells[1]).strip() if cells[1] else ""
            elif a == "Parameter":
                header_row = rn
            elif header_row and a == "Max":
                upper_limit_row = rn
            elif header_row and a == "Min":
                lower_limit_row = rn
            elif lower_limit_row and not data_start_row:
                bin_cell = cells[2] if len(cells) > 2 else None
                if bin_cell is not None:
                    try:
                        int(bin_cell)
                        data_start_row = rn
                    except (ValueError, TypeError):
                        pass

            if data_start_row or rn > 60:
                break

        return meta, header_row, upper_limit_row, lower_limit_row, data_start_row

    def _parse_header(self, ws, header_row: int):
        """Parse header row to discover column positions and param names."""
        row = list(ws.iter_rows(min_row=header_row, max_row=header_row,
                                values_only=True))[0]

        META_COLS = {"Parameter", "Site#"}
        bin_col = x_col = y_col = wafer_id_col = None
        elec_start = None
        param_names: list[str] = []

        for ci, val in enumerate(row):
            if val is None:
                continue
            name = str(val).strip()
            col1 = ci + 1

            if name == "Bin#":
                bin_col = col1
            elif name == "X":
                x_col = col1
            elif name == "Y":
                y_col = col1
            elif name == "ID":
                wafer_id_col = col1
            elif name not in META_COLS:
                if elec_start is None:
                    elec_start = col1
                param_names.append(name)

        return (param_names, bin_col or 3, x_col or 4, y_col or 5,
                wafer_id_col, elec_start or 6)

    def _new_preview(self, ws) -> dict:
        meta, header_row, _ul, _ll, data_start = self._scan_structure(ws)

        if not header_row:
            return {"wafersDetected": 0, "dataRows": 0, "format": "XRW",
                    "productId": meta.get("device_name"),
                    "lotId": meta.get("lot_id"), "paramNames": []}

        param_names, bin_col, _, _, wid_col, _ = self._parse_header(ws, header_row)

        if not data_start:
            return {"wafersDetected": 0, "diePerWafer": None, "dataRows": 0,
                    "format": "XRW", "productId": meta.get("device_name"),
                    "lotId": meta.get("lot_id"), "paramNames": param_names}

        wafer_ids: set[str] = set()
        row_count = 0
        for cells in ws.iter_rows(min_row=data_start, values_only=True):
            bin_v = cells[bin_col - 1] if len(cells) >= bin_col else None
            if bin_v is None:
                break
            try:
                int(bin_v)
            except (ValueError, TypeError):
                break
            row_count += 1
            if wid_col and len(cells) >= wid_col and cells[wid_col - 1]:
                wafer_ids.add(str(cells[wid_col - 1]).strip())

        return {
            "wafersDetected": len(wafer_ids) or 1,
            "diePerWafer": None,
            "dataRows": row_count,
            "format": "XRW",
            "productId": meta.get("device_name"),
            "lotId": meta.get("lot_id"),
            "paramNames": param_names,
        }

    def _new_parse(self, ws) -> ParseResult:
        meta, header_row, ul_row, ll_row, data_start = self._scan_structure(ws)

        empty = ParseResult(product_id="", lot_id="", mark_lot_id=None,
                            test_program=None, vendor_code="XRW",
                            wafers=[], cp_specs=[], param_names=[], total_rows=0)
        if not header_row:
            return empty

        (param_names, bin_col, x_col, y_col,
         wid_col, elec_start) = self._parse_header(ws, header_row)

        # CP specs
        cp_specs: list[ParsedCpSpec] = []
        for i, pname in enumerate(param_names):
            ecol = elec_start + i
            lower = ws.cell(row=ll_row, column=ecol).value if ll_row else None
            upper = ws.cell(row=ul_row, column=ecol).value if ul_row else None
            cp_specs.append(ParsedCpSpec(
                param_name=pname,
                lower_limit=self._safe_float(lower),
                upper_limit=self._safe_float(upper),
            ))

        if not data_start:
            return ParseResult(
                product_id=meta.get("device_name", ""),
                lot_id=meta.get("lot_id", ""),
                mark_lot_id=None,
                test_program=meta.get("test_program"),
                vendor_code="XRW",
                wafers=[], cp_specs=cp_specs,
                param_names=param_names, total_rows=0)

        # Die data
        wafers_dict: dict[str, list[ParsedDie]] = {}
        total_rows = 0
        fallback_wid = meta.get("lot_id", "unknown")

        for row in ws.iter_rows(min_row=data_start, values_only=False):
            bin_val = row[bin_col - 1].value if len(row) >= bin_col else None
            if bin_val is None:
                break
            try:
                bin_int = int(bin_val)
            except (ValueError, TypeError):
                break

            wid_raw = (row[wid_col - 1].value
                       if wid_col and len(row) >= wid_col else None)
            wid = str(wid_raw).strip() if wid_raw else fallback_wid

            x_val = row[x_col - 1].value if len(row) >= x_col else None
            y_val = row[y_col - 1].value if len(row) >= y_col else None

            electrical: dict[str, float | None] = {}
            for j, pname in enumerate(param_names):
                idx = elec_start - 1 + j
                if idx < len(row):
                    electrical[pname] = self._safe_float(row[idx].value)
                else:
                    electrical[pname] = None

            die = ParsedDie(
                site_no=None,
                bin=bin_int,
                x_coord=int(x_val) if x_val is not None else None,
                y_coord=int(y_val) if y_val is not None else None,
                electrical=electrical,
            )
            total_rows += 1
            wafers_dict.setdefault(wid, []).append(die)

        wafers = []
        for wid, dies in wafers_dict.items():
            bin1_count = sum(1 for d in dies if d.bin == 1)
            wafers.append(ParsedWafer(
                wafer_id=wid,
                gross_die=len(dies),
                bin1_count=bin1_count,
                dies=dies,
            ))

        return ParseResult(
            product_id=meta.get("device_name", ""),
            lot_id=meta.get("lot_id", ""),
            mark_lot_id=None,
            test_program=meta.get("test_program"),
            vendor_code="XRW",
            wafers=wafers,
            cp_specs=cp_specs,
            param_names=param_names,
            total_rows=total_rows,
        )

    # ==================================================================
    # Public API — dispatches to legacy or new format
    # ==================================================================

    def preview(self, filepath: str) -> dict:
        wb = self._open_workbook(filepath)
        legacy = self._is_legacy_format(wb)
        ws = self._get_sheet(wb, legacy)
        result = self._legacy_preview(ws) if legacy else self._new_preview(ws)
        wb.close()
        return result

    def parse(self, filepath: str) -> ParseResult:
        wb = self._open_workbook(filepath)
        legacy = self._is_legacy_format(wb)
        ws = self._get_sheet(wb, legacy)
        result = self._legacy_parse(ws) if legacy else self._new_parse(ws)
        wb.close()
        return result
