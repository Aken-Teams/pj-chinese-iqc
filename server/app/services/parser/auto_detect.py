import openpyxl

from app.services.parser.jjw_parser import JJWParser
from app.services.parser.xrw_parser import XRWParser
from app.services.parser.base import BaseParser


def auto_detect_parser(filepath: str) -> BaseParser | None:
    """Heuristic: check row structure to determine format."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        # JJW: header at row 8, check for typical JJW column structure
        # Has data starting at col O (15) for electrical
        row8_val = ws.cell(row=8, column=15).value
        row5_val = ws.cell(row=5, column=12).value

        wb.close()

        # JJW header row 8 col 15 must be a string param name (not a numeric data value)
        if row8_val is not None and isinstance(row8_val, str) and row8_val.strip():
            return JJWParser()

        # XRW header row 5 col 12 must be a string param name
        if row5_val is not None and isinstance(row5_val, str) and row5_val.strip():
            return XRWParser()

    except Exception:
        pass

    return None
