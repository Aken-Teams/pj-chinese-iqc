"""
Dynamic parser that uses VendorFormat column mappings from the database.

This allows any vendor to be parsed as long as their format is configured
in the 廠商管理 (Vendor Management) page.
"""

import openpyxl
from .base import BaseParser, ParseResult, ParsedWafer, ParsedDie, ParsedCpSpec


class DynamicParser(BaseParser):
    """Parser driven by a VendorFormat configuration row."""

    def __init__(self, vendor_code: str, header_row: int, data_start_row: int,
                 lower_limit_row: int, upper_limit_row: int,
                 electrical_start_col: int, wafer_id_col: int, bin_col: int,
                 x_coord_col: int | None = None, y_coord_col: int | None = None,
                 product_id_col: int | None = None, lot_id_col: int | None = None,
                 fixed_die_count: int | None = None,
                 product_id_cell: str | None = None,
                 lot_id_cell: str | None = None):
        self.vendor_code = vendor_code
        self.HEADER_ROW = header_row
        self.DATA_START_ROW = data_start_row
        self.LOWER_LIMIT_ROW = lower_limit_row
        self.UPPER_LIMIT_ROW = upper_limit_row
        self.ELECTRICAL_START_COL = electrical_start_col
        self.WAFER_ID_COL = wafer_id_col
        self.BIN_COL = bin_col
        self.X_COORD_COL = x_coord_col
        self.Y_COORD_COL = y_coord_col
        self.PRODUCT_ID_COL = product_id_col
        self.LOT_ID_COL = lot_id_col
        self.FIXED_DIE_COUNT = fixed_die_count
        self.PRODUCT_ID_CELL = self._parse_cell_ref(product_id_cell)
        self.LOT_ID_CELL = self._parse_cell_ref(lot_id_cell)

    @staticmethod
    def _parse_cell_ref(cell_str: str | None) -> tuple[int, int] | None:
        """Parse a "row,col" string (1-indexed) into a (row, col) tuple."""
        if not cell_str:
            return None
        parts = cell_str.split(",")
        if len(parts) != 2:
            return None
        try:
            return int(parts[0].strip()), int(parts[1].strip())
        except (ValueError, TypeError):
            return None

    @classmethod
    def from_vendor_format(cls, vendor_code: str, fmt) -> "DynamicParser":
        """Create a DynamicParser from a VendorFormat ORM object."""
        return cls(
            vendor_code=vendor_code,
            header_row=fmt.header_row,
            data_start_row=fmt.data_start_row,
            lower_limit_row=fmt.lower_limit_row,
            upper_limit_row=fmt.upper_limit_row,
            electrical_start_col=fmt.electrical_start_col,
            wafer_id_col=fmt.wafer_id_col,
            bin_col=fmt.bin_col,
            x_coord_col=fmt.x_coord_col,
            y_coord_col=fmt.y_coord_col,
            product_id_col=fmt.product_id_col,
            lot_id_col=fmt.lot_id_col,
            fixed_die_count=fmt.fixed_die_count,
            product_id_cell=getattr(fmt, 'product_id_cell', None),
            lot_id_cell=getattr(fmt, 'lot_id_cell', None),
        )

    @staticmethod
    def _safe_float(val):
        if val is None or val == "":
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def _open_workbook(self, filepath: str):
        return openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    def _get_data_sheet(self, wb):
        if "data" in wb.sheetnames:
            return wb["data"]
        return wb[wb.sheetnames[0]]

    def preview(self, filepath: str) -> dict:
        wb = self._open_workbook(filepath)
        ws = self._get_data_sheet(wb)

        # Read param names from header row
        param_names = []
        col = self.ELECTRICAL_START_COL
        while True:
            val = ws.cell(row=self.HEADER_ROW, column=col).value
            if val is None:
                break
            param_names.append(str(val).strip())
            col += 1

        # Count data rows and detect wafers
        wafer_ids = set()
        row_count = 0
        for r in ws.iter_rows(min_row=self.DATA_START_ROW, max_col=self.WAFER_ID_COL, values_only=False):
            cell_val = r[self.WAFER_ID_COL - 1].value
            if cell_val is None:
                break
            wafer_ids.add(str(cell_val))
            row_count += 1

        product_id = None
        lot_id = None
        # Prefer metadata cell over data-row column
        if self.PRODUCT_ID_CELL:
            product_id = ws.cell(row=self.PRODUCT_ID_CELL[0], column=self.PRODUCT_ID_CELL[1]).value
        elif self.PRODUCT_ID_COL:
            product_id = ws.cell(row=self.DATA_START_ROW, column=self.PRODUCT_ID_COL).value
        if self.LOT_ID_CELL:
            lot_id = ws.cell(row=self.LOT_ID_CELL[0], column=self.LOT_ID_CELL[1]).value
        elif self.LOT_ID_COL:
            lot_id = ws.cell(row=self.DATA_START_ROW, column=self.LOT_ID_COL).value

        wb.close()
        return {
            "wafersDetected": len(wafer_ids),
            "diePerWafer": self.FIXED_DIE_COUNT,
            "dataRows": row_count,
            "format": self.vendor_code,
            "productId": str(product_id).strip() if product_id else None,
            "lotId": str(lot_id).strip() if lot_id else None,
            "paramNames": param_names,
        }

    def parse(self, filepath: str) -> ParseResult:
        wb = self._open_workbook(filepath)
        ws = self._get_data_sheet(wb)

        # 1. Read param names
        param_names = []
        col = self.ELECTRICAL_START_COL
        while True:
            val = ws.cell(row=self.HEADER_ROW, column=col).value
            if val is None:
                break
            param_names.append(str(val).strip())
            col += 1

        # 2. Read CP specs
        cp_specs = []
        for i, pname in enumerate(param_names):
            ecol = self.ELECTRICAL_START_COL + i
            lower = ws.cell(row=self.LOWER_LIMIT_ROW, column=ecol).value
            upper = ws.cell(row=self.UPPER_LIMIT_ROW, column=ecol).value
            cp_specs.append(ParsedCpSpec(
                param_name=pname,
                lower_limit=self._safe_float(lower),
                upper_limit=self._safe_float(upper),
            ))

        # 3. Read product/lot from metadata cells if configured
        product_id = None
        lot_id = None
        if self.PRODUCT_ID_CELL:
            v = ws.cell(row=self.PRODUCT_ID_CELL[0], column=self.PRODUCT_ID_CELL[1]).value
            product_id = str(v).strip() if v else ""
        if self.LOT_ID_CELL:
            v = ws.cell(row=self.LOT_ID_CELL[0], column=self.LOT_ID_CELL[1]).value
            lot_id = str(v).strip() if v else ""

        # 4. Read die data, group by wafer
        wafers_dict: dict[str, list[ParsedDie]] = {}
        mark_lot_id = None
        test_program = None
        total_rows = 0

        for row in ws.iter_rows(min_row=self.DATA_START_ROW, values_only=False):
            wafer_val = row[self.WAFER_ID_COL - 1].value
            if wafer_val is None:
                break

            total_rows += 1
            wid = str(wafer_val).strip()

            if product_id is None:
                if self.PRODUCT_ID_COL and self.PRODUCT_ID_COL - 1 < len(row):
                    product_id = str(row[self.PRODUCT_ID_COL - 1].value or "")
                else:
                    product_id = ""
            if lot_id is None:
                if self.LOT_ID_COL and self.LOT_ID_COL - 1 < len(row):
                    lot_id = str(row[self.LOT_ID_COL - 1].value or "")
                else:
                    lot_id = ""
            if mark_lot_id is None:
                mark_lot_id = str(row[2].value or "") if len(row) > 2 else None
                test_program = str(row[4].value or "") if len(row) > 4 else None

            bin_val = row[self.BIN_COL - 1].value
            x_val = row[self.X_COORD_COL - 1].value if self.X_COORD_COL and len(row) >= self.X_COORD_COL else None
            y_val = row[self.Y_COORD_COL - 1].value if self.Y_COORD_COL and len(row) >= self.Y_COORD_COL else None

            electrical = {}
            for j, pname in enumerate(param_names):
                ecol_idx = self.ELECTRICAL_START_COL - 1 + j
                if ecol_idx < len(row):
                    v = row[ecol_idx].value
                    electrical[pname] = self._safe_float(v)
                else:
                    electrical[pname] = None

            die = ParsedDie(
                site_no=None,
                bin=int(bin_val) if bin_val is not None else 0,
                x_coord=int(x_val) if x_val is not None else None,
                y_coord=int(y_val) if y_val is not None else None,
                electrical=electrical,
            )

            if wid not in wafers_dict:
                wafers_dict[wid] = []
            wafers_dict[wid].append(die)

        wb.close()

        # Build wafer objects
        wafers = []
        for wid, dies in wafers_dict.items():
            bin1_count = sum(1 for d in dies if d.bin == 1)
            wafers.append(ParsedWafer(
                wafer_id=wid,
                gross_die=len(dies),
                bin1_count=bin1_count,
                dies=dies,
            ))

        return ParseResult(
            product_id=product_id or "",
            lot_id=lot_id or "",
            mark_lot_id=mark_lot_id,
            test_program=test_program,
            vendor_code=self.vendor_code,
            wafers=wafers,
            cp_specs=cp_specs,
            param_names=param_names,
            total_rows=total_rows,
        )
