"""Build the review-rule spreadsheet a site fills in and sends back.

There is one template, modelled on 徐州's 晶片CP data審核管控標準: a sheet per
vendor, one row per product, and each electrical item spanning Q1/Q2/Q3 with an
L and U limit each. `rules_import_parser` reads exactly this shape, so what goes
out comes back in without translation.

Two things are filled in before the file leaves, and they are the point of the
exercise:

* The parameter rows carry the names the vendor's own CP files use — VTH1,
  VTH2, VTH3 as separate columns, not one generic "VTH". 無錫's own sheet named
  items generically and only 4 of 30 rows could be matched to a real parameter;
  naming them here removes the guesswork at source.

* Q1 is pre-filled from the CP files. Q1 is defined as the vendor's own CP
  limits, so the system already knows it — 9 of 10 stored Q1 rules match their
  CP file exactly. That leaves the site with only Q2 (and optionally Q3) to
  enter, which is the part only they know.
"""

from __future__ import annotations

import io
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.lot import Lot
from app.models.product import Product
from app.models.review import ReviewRule, RuleRevision
from app.models.spec import CpSpec
from app.models.vendor import Vendor

# Q1 | Q2 | Q3, each with a lower and an upper limit.
Q_LEVELS = ("Q1", "Q2", "Q3")
COLS_PER_PARAM = len(Q_LEVELS) * 2
FIRST_DATA_COL = 4          # column D, matching the importer

HEADER_NOTE = (
    "電性管控項目：\n"
    "Q1 L&U Limit 與供應商 CP data 管控一致（系統已依 CP 檔案帶出，可修改）\n"
    "Q2 L&U Limit 由本廠自行制定，看管控項目的良率\n"
    "Q3 選填"
)

_HEAD_FILL = PatternFill("solid", fgColor="EFE9E1")
_LEVEL_FILL = PatternFill("solid", fgColor="F7F4F0")
_BORDER = Border(*[Side(style="thin", color="C9C1B6")] * 4)


def current_version(db: Session, domain: Optional[str]) -> int:
    """The site's ruleset version — 0 until the first import."""
    latest = (db.query(RuleRevision)
              .filter(RuleRevision.domain == domain)
              .order_by(RuleRevision.version.desc())
              .first())
    return latest.version if latest else 0


def _params_for_product(db: Session, product: Product) -> list[str]:
    """Parameter names as the vendor's CP files spell them.

    Taken from the most recent lot: a fab that renames a parameter between
    programs should have the sheet follow the newer file, not the oldest one.
    """
    lot = (db.query(Lot)
           .filter(Lot.product_id == product.id)
           .order_by(Lot.upload_time.desc(), Lot.id.desc())
           .first())
    if not lot:
        return []
    seen, names = set(), []
    for spec in db.query(CpSpec).filter(CpSpec.lot_id == lot.id).all():
        if spec.param_name not in seen:
            seen.add(spec.param_name)
            names.append(spec.param_name)
    return names


def _cp_limits(db: Session, product: Product) -> dict[str, tuple]:
    lot = (db.query(Lot)
           .filter(Lot.product_id == product.id)
           .order_by(Lot.upload_time.desc(), Lot.id.desc())
           .first())
    if not lot:
        return {}
    return {s.param_name: (s.lower_limit, s.upper_limit)
            for s in db.query(CpSpec).filter(CpSpec.lot_id == lot.id).all()}


def build_rules_workbook(db: Session, domain: Optional[str]) -> tuple[bytes, int]:
    """Return (xlsx bytes, version) for one site.

    Only products that actually carry a lot are written. A row for a product
    with no data would be a limit nothing can ever be checked against, which is
    how 無錫 ended up holding 32 rules for 捷捷微 products it has no files for.
    """
    version = current_version(db, domain)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    vendors = db.query(Vendor).order_by(Vendor.code).all()
    wrote_any = False

    for vendor in vendors:
        products = [
            p for p in db.query(Product)
            .filter(Product.vendor_id == vendor.id, Product.domain == domain)
            .order_by(Product.product_code).all()
            if db.query(Lot).filter(Lot.product_id == p.id).count() > 0
        ]
        products = [p for p in products if _params_for_product(db, p)]
        if not products:
            continue

        # Every product's parameters, in first-seen order, so one row per
        # product lines up under a shared set of columns.
        params: list[str] = []
        for p in products:
            for name in _params_for_product(db, p):
                if name not in params:
                    params.append(name)

        ws = wb.create_sheet(vendor.code)
        ws.cell(1, 1, "產品型號")
        ws.cell(1, 2, "%s Product ID" % vendor.code)
        ws.cell(1, 3, "%s 對外型號" % vendor.code)
        note = ws.cell(1, FIRST_DATA_COL, HEADER_NOTE)
        note.alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, 4):
            ws.cell(1, col).font = Font(bold=True)

        for i, name in enumerate(params):
            base = FIRST_DATA_COL + i * COLS_PER_PARAM
            cell = ws.cell(2, base, name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=2, start_column=base,
                           end_row=2, end_column=base + COLS_PER_PARAM - 1)
            for j, level in enumerate(Q_LEVELS):
                lc = base + j * 2
                lv = ws.cell(3, lc, level)
                lv.font = Font(bold=True)
                lv.alignment = Alignment(horizontal="center")
                lv.fill = _LEVEL_FILL
                ws.merge_cells(start_row=3, start_column=lc,
                               end_row=3, end_column=lc + 1)
                for k, label in enumerate(("L LIMIT", "U LIMIT")):
                    h = ws.cell(4, lc + k, label)
                    h.font = Font(size=9)
                    h.fill = _HEAD_FILL
                    h.alignment = Alignment(horizontal="center")

        for r, product in enumerate(products, start=5):
            # Column A is only a label; the importer keys on column B.
            ws.cell(r, 1, product.product_code)
            ws.cell(r, 2, product.product_code).font = Font(bold=True)
            rules = {rule.param_name: rule for rule in db.query(ReviewRule)
                     .filter(ReviewRule.product_id == product.id).all()}
            cp = _cp_limits(db, product)
            own = set(_params_for_product(db, product))
            for i, name in enumerate(params):
                base = FIRST_DATA_COL + i * COLS_PER_PARAM
                if name not in own:
                    # This vendor has the parameter, this product does not.
                    ws.cell(r, base, "/")
                    continue
                rule = rules.get(name)
                q1 = ((rule.q1_lower, rule.q1_upper) if rule
                      and (rule.q1_lower is not None or rule.q1_upper is not None)
                      else cp.get(name, (None, None)))
                pairs = [q1,
                         (rule.q2_lower, rule.q2_upper) if rule else (None, None),
                         (rule.q3_lower, rule.q3_upper) if rule else (None, None)]
                for j, (lo, hi) in enumerate(pairs):
                    for k, v in enumerate((lo, hi)):
                        c = ws.cell(r, base + j * 2 + k,
                                    float(v) if v is not None else None)
                        c.border = _BORDER

        ws.freeze_panes = "D5"
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.row_dimensions[1].height = 62
        for i in range(len(params) * COLS_PER_PARAM):
            ws.column_dimensions[get_column_letter(FIRST_DATA_COL + i)].width = 13
        wrote_any = True

    if not wrote_any:
        ws = wb.create_sheet("EMPTY")
        ws.cell(1, 1, "此廠區目前沒有任何有資料的產品，請先上傳 CP 檔案。")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), version
