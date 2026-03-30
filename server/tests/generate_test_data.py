"""
Generate synthetic Excel CP data files for DZ, VIS, HJM vendors.

Each file follows a different column layout matching a plausible VendorFormat
configuration, so we can test DynamicParser end-to-end.
"""
import os
import random
import openpyxl

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data", "IQC", "extracted")

random.seed(42)

# ---------------------------------------------------------------------------
# Vendor format definitions — these mirror what would be in the DB
# ---------------------------------------------------------------------------

VENDOR_CONFIGS = {
    "DZ": {
        "format_name": "DZ Standard",
        "header_row": 1,
        "data_start_row": 4,
        "lower_limit_row": 2,
        "upper_limit_row": 3,
        "electrical_start_col": 5,
        "wafer_id_col": 1,
        "bin_col": 2,
        "x_coord_col": 3,
        "y_coord_col": 4,
        "product_id_col": None,
        "lot_id_col": None,
        "fixed_die_count": None,
        # Test data generation params
        "product_id": "DZ-P100",
        "lot_id": "DZ-LOT001",
        "num_wafers": 5,
        "dies_per_wafer": 50,
        "params": ["VTH_N", "VTH_P", "IDSAT_N", "IDSAT_P", "IOFF_N"],
        "specs": {
            "VTH_N":   {"lower": 0.35, "upper": 0.55},
            "VTH_P":   {"lower": -0.55, "upper": -0.35},
            "IDSAT_N": {"lower": 400.0, "upper": 700.0},
            "IDSAT_P": {"lower": 180.0, "upper": 350.0},
            "IOFF_N":  {"lower": None, "upper": 1.0e-9},
        },
    },
    "VIS": {
        "format_name": "VIS Format A",
        "header_row": 2,
        "data_start_row": 5,
        "lower_limit_row": 3,
        "upper_limit_row": 4,
        "electrical_start_col": 7,
        "wafer_id_col": 1,
        "bin_col": 3,
        "x_coord_col": 4,
        "y_coord_col": 5,
        "product_id_col": 2,
        "lot_id_col": 6,
        "fixed_die_count": 100,
        # Test data generation params
        "product_id": "VIS-2024A",
        "lot_id": "VIS-L5678",
        "num_wafers": 4,
        "dies_per_wafer": 100,
        "params": ["BV_DSS", "RDS_ON", "VGS_TH", "IGSS", "IDSS", "VF", "GAIN"],
        "specs": {
            "BV_DSS":  {"lower": 55.0, "upper": 80.0},
            "RDS_ON":  {"lower": None, "upper": 0.05},
            "VGS_TH":  {"lower": 1.0, "upper": 3.0},
            "IGSS":    {"lower": None, "upper": 1.0e-7},
            "IDSS":    {"lower": None, "upper": 1.0e-6},
            "VF":      {"lower": 0.3, "upper": 0.9},
            "GAIN":    {"lower": 20.0, "upper": None},
        },
    },
    "HJM": {
        "format_name": "HJM CP Format",
        "header_row": 1,
        "data_start_row": 4,
        "lower_limit_row": 2,
        "upper_limit_row": 3,
        "electrical_start_col": 4,
        "wafer_id_col": 1,
        "bin_col": 2,
        "x_coord_col": None,
        "y_coord_col": None,
        "product_id_col": 3,
        "lot_id_col": None,
        "fixed_die_count": None,
        # Test data generation params
        "product_id": "HJM-3080",
        "lot_id": "HJM-B0012",
        "num_wafers": 3,
        "dies_per_wafer": 80,
        "params": ["IL_1310", "IL_1550", "RL_1310", "RL_1550", "PDL", "WDL"],
        "specs": {
            "IL_1310": {"lower": None, "upper": 0.5},
            "IL_1550": {"lower": None, "upper": 0.7},
            "RL_1310": {"lower": 45.0, "upper": None},
            "RL_1550": {"lower": 40.0, "upper": None},
            "PDL":     {"lower": None, "upper": 0.3},
            "WDL":     {"lower": None, "upper": 0.2},
        },
    },
}


def _gen_value(lower, upper):
    """Generate a random value mostly within spec, with ~5% out-of-spec."""
    if lower is not None and upper is not None:
        mid = (lower + upper) / 2
        spread = (upper - lower) / 2
        # 95% within spec, 5% outside
        if random.random() < 0.05:
            return mid + random.uniform(-spread * 2, spread * 2)
        return random.gauss(mid, spread * 0.3)
    elif lower is not None:
        # lower-only: most values well above lower
        return lower + abs(random.gauss(0, lower * 0.1 if lower != 0 else 1.0))
    elif upper is not None:
        # upper-only: most values well below upper
        return upper * random.uniform(0.1, 0.95)
    return random.gauss(0, 1)


def generate_vendor_file(vendor_code: str, config: dict) -> str:
    """Generate an xlsx file for this vendor and return the file path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"

    header_row = config["header_row"]
    data_start = config["data_start_row"]
    limit_row = config["lower_limit_row"]   # we put both lower & upper here
    ecol_start = config["electrical_start_col"]
    wid_col = config["wafer_id_col"]
    bin_col = config["bin_col"]
    x_col = config.get("x_coord_col")
    y_col = config.get("y_coord_col")
    pid_col = config.get("product_id_col")
    lid_col = config.get("lot_id_col")
    params = config["params"]
    specs = config["specs"]

    # -- Header row: write param names in electrical columns
    for i, pname in enumerate(params):
        ws.cell(row=header_row, column=ecol_start + i, value=pname)

    # Also write column labels for non-electrical columns in header
    ws.cell(row=header_row, column=wid_col, value="Wafer_ID")
    ws.cell(row=header_row, column=bin_col, value="Bin")
    if x_col:
        ws.cell(row=header_row, column=x_col, value="X")
    if y_col:
        ws.cell(row=header_row, column=y_col, value="Y")
    if pid_col:
        ws.cell(row=header_row, column=pid_col, value="Product_ID")
    if lid_col:
        ws.cell(row=header_row, column=lid_col, value="Lot_ID")

    # -- Limit row: lower limits on limit_row, upper limits also on limit_row
    # For DynamicParser: lower_limit_row reads lower, upper_limit_row reads upper
    # Since we set both to the same row for DZ/HJM, we store lower limit values.
    # Actually, the parser reads DIFFERENT rows for lower vs upper.
    # Let's use separate rows when they differ.

    lower_row = config["lower_limit_row"]
    upper_row = config["upper_limit_row"]

    for i, pname in enumerate(params):
        spec = specs[pname]
        col = ecol_start + i
        ws.cell(row=lower_row, column=col, value=spec["lower"])
        ws.cell(row=upper_row, column=col, value=spec["upper"])

    # -- Data rows
    current_row = data_start
    for w in range(config["num_wafers"]):
        wafer_id = f"W{w + 1:02d}"
        for d in range(config["dies_per_wafer"]):
            ws.cell(row=current_row, column=wid_col, value=wafer_id)

            # Generate electrical values and determine bin
            values = {}
            all_pass = True
            for pname in params:
                spec = specs[pname]
                val = _gen_value(spec["lower"], spec["upper"])
                values[pname] = val
                lo = spec["lower"]
                hi = spec["upper"]
                if lo is not None and val <= lo:
                    all_pass = False
                if hi is not None and val >= hi:
                    all_pass = False

            bin_val = 1 if all_pass else random.choice([2, 3, 4])
            ws.cell(row=current_row, column=bin_col, value=bin_val)

            if x_col:
                ws.cell(row=current_row, column=x_col, value=d % 10)
            if y_col:
                ws.cell(row=current_row, column=y_col, value=d // 10)
            if pid_col:
                ws.cell(row=current_row, column=pid_col, value=config["product_id"])
            if lid_col:
                ws.cell(row=current_row, column=lid_col, value=config["lot_id"])

            for i, pname in enumerate(params):
                ws.cell(row=current_row, column=ecol_start + i, value=round(values[pname], 8))

            current_row += 1

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"test_{vendor_code}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    wb.close()
    print(f"  Generated {filepath}  ({config['num_wafers']} wafers × {config['dies_per_wafer']} dies)")
    return filepath


def main():
    print("Generating synthetic CP data files...")
    paths = {}
    for vendor_code, config in VENDOR_CONFIGS.items():
        paths[vendor_code] = generate_vendor_file(vendor_code, config)
    print(f"\nDone. Files: {list(paths.values())}")
    return paths


if __name__ == "__main__":
    main()
